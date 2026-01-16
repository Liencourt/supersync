from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404, redirect,render
from django.urls import reverse_lazy
from django.views.generic import DetailView
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from .models import Evento,Grade,GradeGrupo, ItemGrade, ItemGradeSKU,ItemGradeDistribuicao
from .forms import GradeForm, ItemGradeForm,EventoForm,GradeHeaderForm
from django.http import JsonResponse
from django.urls import reverse
from django.db import transaction,connection
from usuarios.models import Associado
from django.db.models import Sum, Value, DecimalField,Min, Max, Count, Q
from django.db.models.functions import Coalesce
import json
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from google.cloud import bigquery
import os
from google.oauth2 import service_account # Precisa disso para ler o JSON


try:
    from gcp_services.services import bigquery_client
except ImportError:
    bigquery_client = None

try:
    bq_client = bigquery.Client()
except Exception:
    bq_client = None



# 1. Listagem (READ)
class EventoListView(LoginRequiredMixin, ListView):
    model = Evento
    template_name = 'apuracao_grade/evento_list.html'
    context_object_name = 'eventos'
    paginate_by = 10

    def get_queryset(self):
        # Retorna ordenado pela data de início (mais recente primeiro)
        return Evento.objects.all().order_by('-data_inicio')

# 2. Cadastro (CREATE)
class EventoCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Evento
    form_class = EventoForm
    template_name = 'apuracao_grade/evento_form.html'
    success_url = reverse_lazy('apuracao_grade:evento-list')
    success_message = "Evento '%(descricao)s' criado com sucesso!"

# 3. Edição (UPDATE)
class EventoUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Evento
    form_class = EventoForm
    template_name = 'apuracao_grade/evento_form.html'
    success_url = reverse_lazy('apuracao_grade:evento-list')
    success_message = "Evento atualizado com sucesso!"

# 4. Exclusão (DELETE)
class EventoDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Evento
    template_name = 'apuracao_grade/evento_confirm_delete.html'
    success_url = reverse_lazy('apuracao_grade:evento-list')
    success_message = "Evento excluído com sucesso!"
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, self.success_message)
        return super().delete(request, *args, **kwargs)
    
#grade views.py
try:
    from gcp_services.services import bigquery_client
except ImportError:
    bigquery_client = None

class GradeCreateView(LoginRequiredMixin, CreateView):
    model = Grade
    form_class = GradeForm
    template_name = 'apuracao_grade/grade_form.html'
    success_url = reverse_lazy('apuracao_grade:grade-list')

    def form_valid(self, form):
        # Transaction garante que só salva a Grade se salvar os grupos também
        with transaction.atomic():
            # 1. Salva a Grade (Cabeçalho)
            self.object = form.save()
            
            # 2. Pega o JSON do campo oculto
            grupos_str = form.cleaned_data.get('grupos_json')
            if grupos_str:
                grupos_list = json.loads(grupos_str)
                
                # 3. Cria os registros na tabela filha
                for item in grupos_list:
                    GradeGrupo.objects.create(
                        grade=self.object,
                        grupo_id=item['id'],
                        grupo_nome=item['nome']
                    )
                    
        return super().form_valid(form) # Redireciona

class GradeListView(LoginRequiredMixin, ListView):
    model = Grade
    template_name = 'apuracao_grade/grade_list.html'
    context_object_name = 'grades'
    paginate_by = 10 

    def get_queryset(self):
        # 1. Otimização de queries (evita N+1 no template)
        qs = Grade.objects.select_related('evento').prefetch_related('grupos').order_by('-id')
        
        # 2. Captura o termo de busca da URL
        query = self.request.GET.get('q')
        
        # 3. Se houver busca, filtra
        if query:
            qs = qs.filter(
                Q(id__icontains=query) |  # Busca por ID
                Q(evento__descricao__icontains=query) |  # Busca por nome do evento
                Q(comprador__name__icontains=query) |        # Busca pelo Nome do Usuário vinculado
                Q(comprador__username__icontains=query) |    # (Opcional) Busca pelo Login também
                Q(grupos__grupo_nome__icontains=query)   # Busca por nome do grupo/fornecedor
            ).distinct() # Distinct é vital aqui porque a busca em 'grupos' (ManyToMany) pode duplicar linhas
            
        return qs

    # Opcional: Passar o valor da busca de volta pro template para manter no input
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_term'] = self.request.GET.get('q', '')
        return context
    
def buscar_fornecedores_api(request):
    """
    Versão HÍBRIDA:
    1. Tenta usar Autenticação Automática (Cloud Run).
    2. Se der erro, tenta usar o arquivo 'credenciais.json' (Local).
    """
    termo = request.GET.get('term', '').upper().strip()
    termo_limpo = termo.replace('.', '').replace('/', '').replace('-', '')
    
    if len(termo) < 3:
        return JsonResponse([], safe=False)

    client = None

    # --- BLOCO DE CONEXÃO INTELIGENTE ---
    try:
        # TENTATIVA 1: Modo Nuvem (Automático)
        # O Google tenta achar a credencial sozinho.
        client = bigquery.Client()
    except Exception as e:
        print(f"Modo Nuvem falhou, tentando modo Local... Erro: {e}")
        
        # TENTATIVA 2: Modo Local (Arquivo JSON)
        try:
            # Ajuste o caminho se seu arquivo estiver em outra pasta
            caminho_json = 'credenciais.json' 
            
            if os.path.exists(caminho_json):
                credentials = service_account.Credentials.from_service_account_file(caminho_json)
                client = bigquery.Client(credentials=credentials)
            else:
                return JsonResponse([{'id': 0, 'text': 'Erro: credenciais.json não encontrado'}], safe=False)
        except Exception as e_local:
             return JsonResponse([{'id': 0, 'text': f'Erro Fatal Auth: {str(e_local)}'}], safe=False)

    # --- FIM DO BLOCO DE CONEXÃO ---

    # Query SQL
    tabela_fornecedores = "`singular-ray-422121`.gold.dim_fornecedor" 
    
    sql = f"""
        SELECT 
            SEQREDE as id, 
            NOME_REDE as grupo,
            cnpj_completo as cnpj,
            NOMERAZAO as razao
        FROM {tabela_fornecedores}
        WHERE 
           (
               UPPER(NOME_REDE) LIKE '%{termo}%'
            OR UPPER(NOMERAZAO) LIKE '%{termo}%'
            OR REPLACE(REPLACE(REPLACE(cnpj_completo, '.', ''), '/', ''), '-', '') LIKE '%{termo_limpo}%'
            OR CAST(SEQREDE AS STRING) LIKE '%{termo}%'
           )
        LIMIT 50
    """

    try:
        query_job = client.query(sql)
        dados = query_job.result()
        
        resultados = []
        ids_processados = set()
        
        for linha in dados:
            raw_id = linha.get('id')
            seq = str(raw_id) if raw_id is not None else "0"
            
            if seq in ids_processados:
                continue
            ids_processados.add(seq)

            grupo = linha.get('grupo') or 'SEM NOME'
            razao = linha.get('razao') or ''
            
            texto_exibicao = f"{grupo} | {razao}" if razao else grupo
            
            resultados.append({
                'id': seq,
                'text': texto_exibicao
            })
             
        return JsonResponse(resultados, safe=False)

    except Exception as e:
        return JsonResponse([{'id': 0, 'text': f'Erro SQL: {str(e)}'}], safe=False)
        
class GradeDetalheView(LoginRequiredMixin, DetailView):
    model = Grade
    template_name = 'apuracao_grade/grade_detalhe.html'
    context_object_name = 'grade'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_item'] = ItemGradeForm()
        
        # Busca os itens já somando o que foi distribuído para as lojas
        # Coalesce(..., 0) serve para transformar 'None' em '0' se não tiver distribuição
        itens_anotados = self.object.itens.annotate(
            total_distribuido=Coalesce(
                Sum('distribuicoes__volume_compra_minima'), 
                Value(0),
                output_field=DecimalField()
            )
        ).order_by('id')
        
        context['itens_com_status'] = itens_anotados
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object() # Pega a grade atual pelo ID na URL
        
        form = ItemGradeForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 1. Salva o ItemGrade (Pai)
                    item = form.save(commit=False)
                    item.grade = self.object
                    item.save()
                    
                    # 2. Salva os SKUs (Filhos) vindos do JSON
                    skus_str = form.cleaned_data.get('skus_json')
                    if skus_str:
                        skus_list = json.loads(skus_str)
                        for sku_data in skus_list:
                            # sku_data = {'id': 123, 'text': 'NOME PRODUTO'}
                            ItemGradeSKU.objects.create(
                                item_grade=item,
                                codigo_produto=sku_data['id'],
                                descricao_produto=sku_data['text']
                                # Removemos embalagem daqui pois vem do pai agora
                            )
                
                # Sucesso: Recarrega a página
                return redirect('apuracao_grade:grade-detalhe', pk=self.object.pk)
            
            except Exception as e:
                print(f"Erro ao salvar item: {e}")
                form.add_error(None, f"Erro ao salvar dados: {e}")

        # Se falhar, re-renderiza a página com o modal aberto e erros
        context = self.get_context_data()
        context['form_item'] = form
        context['modal_open'] = True
        return self.render_to_response(context)


def buscar_produtos_api(request):
    """
    API para buscar produtos (SKUs) no BigQuery.
    Versão BLINDADA: Usa autenticação nativa (Cloud Run) ou arquivo local.
    """
    termo = request.GET.get('term', '').upper().strip()
    
    if len(termo) < 3:
        return JsonResponse([], safe=False)
    
    # --- 1. CONEXÃO HÍBRIDA (IGUAL À DE FORNECEDORES) ---
    client = None
    try:
        # Tenta conectar automaticamente (Funciona no Cloud Run)
        client = bigquery.Client()
    except Exception as e:
        # Se falhar, tenta o arquivo local (Funciona no seu PC)
        try:
            caminho_json = 'credenciais.json' 
            if os.path.exists(caminho_json):
                credentials = service_account.Credentials.from_service_account_file(caminho_json)
                client = bigquery.Client(credentials=credentials)
            else:
                print("Erro: credenciais.json não encontrado para produtos")
                return JsonResponse([], safe=False)
        except Exception as e_local:
            print(f"Erro Conexão Produtos: {e_local}")
            return JsonResponse([], safe=False)
    # ----------------------------------------------------

    # Prepara termo para busca (troca espaço por %)
    termo_smart = termo.replace(' ', '%')

    tabela_produtos = "`singular-ray-422121`.landing_saerj.DIM_PRODUTOS" 

    sql = f"""
        SELECT DISTINCT
            CAST(SEQPRODUTO AS INT64) as id,
            DESCCOMPLETA as text
        FROM {tabela_produtos}
        WHERE 
           (
               UPPER(DESCCOMPLETA) LIKE '%{termo_smart}%'
               OR CAST(SEQPRODUTO AS STRING) LIKE '%{termo}%'
               OR CAST(CODACESSO AS STRING) LIKE '%{termo}%'
               OR UPPER(MARCA) LIKE '%{termo_smart}%'
           )
        ORDER BY text
        LIMIT 300
    """
    
    try:
        # Executa a query usando o cliente nativo
        query_job = client.query(sql)
        dados = query_job.result()
        
        resultados = []
        for linha in dados:
        
            try:
                id_limpo = int(linha.get('id'))
            except:
                id_limpo = linha.get('id')

            resultados.append({
                'id': id_limpo, 
                'text': linha.get('text'),
            })
            
        return JsonResponse(resultados, safe=False)

    except Exception as e:
        print(f"Erro API Produtos SQL: {e}")
        # Retorna lista vazia em caso de erro para não quebrar o front
        return JsonResponse([], safe=False)

@require_POST 
def excluir_item_grade(request, pk):
    """
    Exclui um item (agrupamento) da grade e redireciona de volta.
    """
    item = get_object_or_404(ItemGrade, pk=pk)
    grade_id = item.grade.id # Guardamos o ID da Grade Pai para voltar pra ela
    
    try:
        item.delete()
        # Mensagem de sucesso via messages.success(request, ...)
    except Exception as e:
        print(f"Erro ao excluir item: {e}")
    
    return redirect('apuracao_grade:grade-detalhe', pk=grade_id)
    
# 1. Função para buscar Lojas Ativas (Usa o Model Associado que lê do Postgres)
def get_lojas_ativas():
    try:
        # Busca da tabela 'cadastro_associado' via Model
        lojas = Associado.objects.filter(status='ATIVO').order_by('nome')
        return [{'id': loja.id, 'nome': loja.nome} for loja in lojas]
    except Exception as e:
        print(f"Erro ao buscar associados: {e}")
        return []

# 2. Função para buscar Histórico de % (Query SQL Direta)
def get_sugestao_distribuicao():
    """
    Busca o histórico.
    Protegido para não quebrar se a tabela/view não existir no banco.
    """
    sugestao = {}
    
    # Removemos o 'public.' para compatibilidade e adicionamos try/except robusto
    sql = """
        SELECT 
            UPPER(TRIM(nomeassociado)) as nome, 
            SUM(percentual_qtd) as perc
        FROM gradepercatualassoc 
        WHERE (ano, trimestre) IN (
            SELECT ano, MAX(trimestre) 
            FROM gradepercatualassoc 
            WHERE ano = (SELECT MAX(ano) FROM gradepercatualassoc)
            GROUP BY ano
        )
        GROUP BY nomeassociado
    """
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(sql)
            rows = cursor.fetchall()
            
            for row in rows:
                nome = row[0]
                perc = float(row[1]) if row[1] is not None else 0.0
                sugestao[nome] = perc
                
    except Exception as e:
        # Aqui capturamos o erro "relation does not exist" silenciosamente
        # para permitir que a tela de distribuição abra mesmo sem histórico.
        print(f"⚠️ AVISO: Não foi possível carregar histórico de distribuição.")
        print(f"Erro detalhado: {e}")
        # Retorna vazio, assim o sistema assume 0% para todos sem travar
        return {}
    
    return sugestao

# 3. View Principal
def gerenciar_distribuicao(request, pk):
    item = get_object_or_404(ItemGrade, pk=pk)
    
    # --- GET: Carregar dados ---
    if request.method == 'GET':
        # Verifica se o front pediu para RESETAR (Recalcular do zero)
        forcar_reset = request.GET.get('reset') == 'true'
        
        distribuicoes = []
        
        # Se NÃO for reset, tenta buscar o que já está salvo no banco
        if not forcar_reset:
            distribuicoes = list(item.distribuicoes.all().values('associado_id', 'associado_nome', 'percentual_participacao', 'volume_compra_minima'))
        
        # Se não achou nada salvo OU se for um reset forçado, calcula a sugestão
        if not distribuicoes:
            lojas = get_lojas_ativas()
            historico = get_sugestao_distribuicao()
            
            data = []
            vol_total = float(item.volume_negociado)

            for loja in lojas:
                nome_chave = loja['nome'].upper().strip()
                perc = historico.get(nome_chave, 0)
                vol_sugerido = round(vol_total * (perc / 100))
                
                data.append({
                    'associado_id': loja['id'],
                    'associado_nome': loja['nome'],
                    'percentual_participacao': perc,
                    'volume_compra_minima': vol_sugerido
                })
        else:
            data = distribuicoes

        return JsonResponse({
            'item_volume': item.volume_negociado,
            'qtd_embalagem': item.quantidade_embalagem,
            'lojas': data
        }, safe=False)

    # --- POST: Salvar (continua igual) ---
    if request.method == 'POST':
        # ... (seu código de salvar existente não muda nada) ...
        # Vou omitir aqui para economizar espaço, mantenha o que já estava funcionando
        import json
        try:
            dados = json.loads(request.body)
            lojas_data = dados.get('lojas', [])
            qtd_emb = float(item.quantidade_embalagem)
            
            with transaction.atomic():
                item.distribuicoes.all().delete()
                for loja in lojas_data:
                    vol = float(loja.get('volume', 0))
                    perc = float(loja.get('percentual', 0))
                    if vol > 0 or perc > 0:
                        vol_fisico = vol * qtd_emb
                        ItemGradeDistribuicao.objects.create(
                            item_grade=item,
                            associado_id=loja['id'],
                            associado_nome=loja['nome'],
                            percentual_participacao=perc,
                            volume_compra_minima=vol,
                            volume_fisico=vol_fisico
                        )
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'erro', 'message': str(e)}, status=400)
        
# 1. API para buscar os dados do item e preencher o modal
def api_obter_item(request, pk):
    item = get_object_or_404(ItemGrade, pk=pk)
    
    # Prepara a lista de SKUs
    skus = [{'id': s.codigo_produto, 'text': s.descricao_produto} for s in item.skus.all()]
    
    data = {
        'id': item.id,
        'descricao_resumida': item.descricao_resumida,
        'unidade_medida': item.unidade_medida,
        'quantidade_embalagem': item.quantidade_embalagem,
        'volume_negociado': item.volume_negociado,
        'custo_bruto': item.custo_bruto,
        'custo_liquido': item.custo_liquido,
        'verba_sell_in': item.verba_sell_in,
        'desconto_boleto': item.desconto_boleto,
        'skus': skus # Manda a lista para o JS montar as tags
    }
    return JsonResponse(data)

# 2. View para Salvar a Edição (UPDATE)
def editar_item_grade(request, pk):
    item = get_object_or_404(ItemGrade, pk=pk)
    
    if request.method == 'POST':
        form = ItemGradeForm(request.POST, instance=item)
        if form.is_valid():
            with transaction.atomic():
                item = form.save()
                
                # Atualiza os SKUs (Apaga os antigos e cria os novos da lista)
                skus_str = form.cleaned_data.get('skus_json')
                if skus_str:
                    item.skus.all().delete() # Limpa anteriores
                    skus_list = json.loads(skus_str)
                    for sku_data in skus_list:
                        ItemGradeSKU.objects.create(
                            item_grade=item,
                            codigo_produto=sku_data['id'],
                            descricao_produto=sku_data['text']
                        )
            # Redireciona de volta para a grade pai
            return redirect('apuracao_grade:grade-detalhe', pk=item.grade.id)
            
    return redirect('apuracao_grade:grade-detalhe', pk=item.grade.id)

@require_POST
def finalizar_grade(request, pk):
    """
    Valida se todos os itens foram 100% distribuídos e altera o status da grade.
    """
    grade = get_object_or_404(Grade, pk=pk)
    
    # 1. Verifica se tem itens
    if not grade.itens.exists():
        return JsonResponse({'status': 'erro', 'message': 'A grade não possui itens cadastrados.'})

    # 2. Validação Matemática (Item a Item)
    itens_pendentes = []
    
    # Anota cada item com o total distribuído
    itens = grade.itens.annotate(
        total_dist=Coalesce(Sum('distribuicoes__volume_compra_minima'), Value(0), output_field=DecimalField())
    )
    
    for item in itens:
        # Compara com margem de tolerância para float (0.1)
        diff = abs(item.volume_negociado - item.total_dist)
        if diff > 0.1:
            itens_pendentes.append(f"- {item.descricao_resumida} (Diferença: {diff:.0f})")
    
    # 3. Se houver pendências, bloqueia e retorna lista
    if itens_pendentes:
        msg = "\n".join(itens_pendentes)
        return JsonResponse({'status': 'erro', 'message': msg})
    
    # 4. Sucesso: Muda Status
    grade.status = 'concluida' # Ou 'aprovada', conforme seu fluxo
    grade.save()
    
    return JsonResponse({'status': 'ok', 'message': 'Grade finalizada com sucesso!'})

def exportar_grade_excel(request, pk):
    """
    Gera um arquivo Excel com a matriz de distribuição.
    CORREÇÃO: Conversão de tipos Decimal/Float ajustada.
    """
    grade = get_object_or_404(Grade, pk=pk)
    
    # --- 1. CONFIGURAÇÃO VISUAL ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Grade {grade.id}"
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- 2. CABEÇALHO DO ARQUIVO (LINHAS 1-5) ---
    ws['A1'] = "ESPELHO DE DISTRIBUIÇÃO - SUPER SYNC"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Grade: #{grade.id}"
    ws['A3'] = f"Evento: {grade.evento.descricao}"
    nome_comprador = ""
    if grade.comprador:
        nome_comprador = grade.comprador.name or grade.comprador.username

    ws['A4'] = f"Comprador: {nome_comprador}"
    
    # ADICIONADO: CAMPO OBSERVAÇÃO
    ws['A5'] = f"Obs: {grade.observacoes or 'Nenhuma observação registrada.'}"
    ws.merge_cells('A5:H5') # Mescla para caber texto longo
    ws['A5'].font = Font(italic=True, color="555555")

    ws['D2'] = f"Status: {grade.get_status_display()}"
    ws['D3'] = f"Vigência: {grade.data_inicio.strftime('%d/%m/%Y')} a {grade.data_fim.strftime('%d/%m/%Y')}"

    # --- 3. ANÁLISE INTELIGENTE DE COLUNAS (%) ---
    # Vamos descobrir se cada loja tem % fixo ou variável para decidir se cria 1 ou 2 colunas
    
    total_itens = grade.itens.count()
    lojas = Associado.objects.filter(status='ATIVO').order_by('nome')
    
    # Agrega estatísticas da distribuição para saber se varia
    stats = ItemGradeDistribuicao.objects.filter(item_grade__grade=grade).values('associado_id').annotate(
        min_p=Min('percentual_participacao'),
        max_p=Max('percentual_participacao'),
        qtd_registros=Count('id')
    )
    # Transforma em dicionário para acesso rápido: { loja_id: {min, max, count} }
    stats_dict = {s['associado_id']: s for s in stats}
    
    colunas_lojas = [] # Vai guardar a definição: (loja_obj, 'simples' ou 'dupla', valor_percentual)
    
    for loja in lojas:
        stat = stats_dict.get(loja.id)
        
        is_constante = False
        perc_fixo = 0
        
        if stat:
            # Regra: Tem registro em TODOS os itens E (Min == Max)
            if stat['qtd_registros'] == total_itens and stat['min_p'] == stat['max_p']:
                is_constante = True
                perc_fixo = stat['min_p']
            # Se não tem registros (0%), tecnicamente é constante, mas vamos tratar como variável se tiver algum
            elif stat['qtd_registros'] == 0:
                is_constante = True # Constante em 0%
                perc_fixo = 0
        else:
            # Loja sem nenhuma distribuição salva
            is_constante = True
            perc_fixo = 0
            
        if is_constante:
            colunas_lojas.append({'loja': loja, 'tipo': 'simples', 'perc': perc_fixo})
        else:
            colunas_lojas.append({'loja': loja, 'tipo': 'dupla', 'perc': None})

    # --- 4. MONTAR CABEÇALHOS DA TABELA (LINHA 7) ---
    row_header = 7
    titulos_fixos = ["ID", "Descrição Item", "Emb.", "Custo Líq.", "Vol. Negociado", "Vol. Distribuído"]
    
    col_idx = 1
    # Títulos Fixos
    for titulo in titulos_fixos:
        cell = ws.cell(row=row_header, column=col_idx, value=titulo)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        col_idx += 1
        
    # Títulos Dinâmicos (Lojas)
    loja_col_map = {} # Mapa para saber onde começa cada loja: {loja_id: col_inicial}
    
    for info in colunas_lojas:
        loja = info['loja']
        loja_col_map[loja.id] = col_idx # Salva onde começa
        
        if info['tipo'] == 'simples':
            # Apenas 1 coluna: "Nome (X%)"
            texto = f"{loja.nome}\n({info['perc']:.2f}%)"
            cell = ws.cell(row=row_header, column=col_idx, value=texto)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center') # Wrap para quebrar linha
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            col_idx += 1
            
        else:
            # 2 Colunas: "Nome (Vol)" e "Nome (%)"
            # Mescla o nome da loja em cima
            ws.merge_cells(start_row=row_header-1, start_column=col_idx, end_row=row_header-1, end_column=col_idx+1)
            topo = ws.cell(row=row_header-1, column=col_idx, value=loja.nome)
            topo.alignment = center_align
            topo.font = bold_font
            topo.fill = header_fill
            
            # Sub-cabeçalhos
            c1 = ws.cell(row=row_header, column=col_idx, value="Vol.")
            c2 = ws.cell(row=row_header, column=col_idx+1, value="%")
            
            for c in [c1, c2]:
                c.font = bold_font
                c.fill = header_fill
                c.alignment = center_align
                c.border = thin_border
            
            col_idx += 2

    # Ajuste de larguras fixas
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['D'].width = 12
    
    # --- 5. PREENCHIMENTO DOS DADOS ---
    row_num = 8
    itens = grade.itens.all().prefetch_related('distribuicoes')
    
    for item in itens:
        # Dicionário rápido: { loja_id: {vol, perc} }
        dist_data = {d.associado_id: {'vol': d.volume_compra_minima, 'perc': d.percentual_participacao} for d in item.distribuicoes.all()}
        
        total_dist = sum(d['vol'] for d in dist_data.values())
        
        # Colunas Fixas
        ws.cell(row=row_num, column=1, value=item.id).border = thin_border
        ws.cell(row=row_num, column=2, value=item.descricao_resumida).border = thin_border
        ws.cell(row=row_num, column=3, value=item.get_unidade_medida_display()).alignment = center_align
        
        c_custo = ws.cell(row=row_num, column=4, value=float(item.custo_liquido))
        c_custo.number_format = '#,##0.00'
        c_custo.border = thin_border
        
        ws.cell(row=row_num, column=5, value=float(item.volume_negociado)).border = thin_border
        
        # Validação Total (Vermelho se divergir)
        c_total = ws.cell(row=row_num, column=6, value=total_dist)
        c_total.border = thin_border
        if abs(float(total_dist) - float(item.volume_negociado)) > 0.1:
            c_total.font = Font(color="FF0000", bold=True)
            
        # Colunas Dinâmicas
        for info in colunas_lojas:
            loja = info['loja']
            col_atual = loja_col_map[loja.id]
            
            dados = dist_data.get(loja.id, {'vol': 0, 'perc': 0})
            
            if info['tipo'] == 'simples':
                # Só escreve o Volume
                cell = ws.cell(row=row_num, column=col_atual, value=dados['vol'])
                cell.border = thin_border
                if dados['vol'] > 0: cell.font = Font(bold=True)
                
            else:
                # Escreve Volume e Percentual
                c_vol = ws.cell(row=row_num, column=col_atual, value=dados['vol'])
                c_perc = ws.cell(row=row_num, column=col_atual+1, value=dados['perc']/100) # Divide por 100 para formato %
                
                c_vol.border = thin_border
                c_perc.border = thin_border
                c_perc.number_format = '0.00%' 
                
                if dados['vol'] > 0: 
                    c_vol.font = Font(bold=True)
        
        row_num += 1

    # Finalização
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Grade_{grade.id}_Espelho.xlsx'
    wb.save(response)
    
    return response


def api_dashboard_apuracao(request, pk):
    """
    API do Dashboard usando 'bigquery_client'.
    Retorna dados cruzando META (Django) x REALIZADO (BigQuery).
    """
    if not bigquery_client:
        return JsonResponse({'status': 'erro', 'message': 'Serviço GCP BigQuery não disponível.'}, status=500)

    grade = get_object_or_404(Grade, pk=pk)

    # --- PASSO 0: PREPARAR PARÂMETROS ---
    dt_inicio = grade.data_inicio.strftime('%Y-%m-%d')
    dt_fim = grade.data_fim.strftime('%Y-%m-%d')
    dt_fim_evento = grade.evento.data_fim.strftime('%Y-%m-%d')
    
    # Mapeamento SKU -> Produto Pai
    skus_objs = ItemGradeSKU.objects.filter(item_grade__grade=grade).select_related('item_grade')
    skus_map = {str(s.codigo_produto): s.item_grade.descricao_resumida for s in skus_objs}
    skus_list = list(skus_map.keys())
    
    grupos_ids = [g.grupo_id for g in grade.grupos.all()]

    if not skus_list:
        return JsonResponse({'status': 'erro', 'message': 'Nenhum SKU vinculado à grade.'})
    if not grupos_ids:
        return JsonResponse({'status': 'erro', 'message': 'Nenhum Grupo vinculado.'})

    try:
        # --- PASSO 1: LISTA BRANCA DE CNPJS ---
        grupos_sql = ", ".join(f"'{g}'" for g in grupos_ids)
        query_cnpjs = f"""
            SELECT DISTINCT CAST(cnpj_completo AS STRING) as cnpj
            FROM `singular-ray-422121.gold.dim_fornecedor`
            WHERE CAST(SEQREDE AS STRING) IN ({grupos_sql})
        """
        dados_cnpjs = bigquery_client.run_query(query_cnpjs)
        
        cnpjs_validos = []
        if dados_cnpjs:
            for row in dados_cnpjs:
                val = row.get('cnpj') if isinstance(row, dict) else row[0]
                if val: cnpjs_validos.append(val)

        if not cnpjs_validos:
            cnpjs_validos = ['00000000000000']

        # Formatação Arrays SQL
        skus_array_str = "[" + ", ".join(f"'{s}'" for s in skus_list) + "]"
        cnpjs_array_str = "[" + ", ".join(f"'{c}'" for c in cnpjs_validos) + "]"

        # --- PASSO 2: QUERY GERAL (GRÁFICOS TIMELINE) ---
        sql_apuracao = f"""
            SELECT 
                data_emissao, 
                NomeAssociado, 
                SUM(CASE WHEN CAST(CNPJ_fornecedor AS STRING) IN UNNEST({cnpjs_array_str}) THEN QtdCompra ELSE 0 END) AS qtd_homologada,
                SUM(CASE WHEN CAST(CNPJ_fornecedor AS STRING) NOT IN UNNEST({cnpjs_array_str}) THEN QtdCompra ELSE 0 END) AS qtd_pirata,
                SUM(CASE WHEN data_emissao > '{dt_fim}' THEN QtdCompra ELSE 0 END) AS qtd_atrasada
            FROM `singular-ray-422121.gold.obt_tb_compra_agg`
            WHERE 
                data_emissao BETWEEN '{dt_inicio}' AND '{dt_fim_evento}'
                AND CAST(SeqProduto AS STRING) IN UNNEST({skus_array_str})
            GROUP BY 1, 2
            ORDER BY data_emissao
        """
        print("\n" + "="*50)
        print(">>> SQL apuracao:")
        print(sql_apuracao)
        print("="*50 + "\n")

        res_apuracao = bigquery_client.run_query(sql_apuracao)
        df_apuracao = [dict(row) for row in res_apuracao] if res_apuracao else []

        # --- PASSO 3: QUERY OFENSORES ---
        # Adicionei tratamento extra para garantir nome
        sql_ofensores = f"""
            SELECT 
                COALESCE(f.NOMERAZAO, f.FANTASIA, CONCAT('CNPJ: ', CAST(c.CNPJ_fornecedor AS STRING))) as fornecedor_nome,
                SUM(c.QtdCompra) as volume_pirata,
                SUM(c.ValorCompraBruta) as valor_desviado
            FROM `singular-ray-422121.gold.obt_tb_compra_agg` c
            LEFT JOIN `singular-ray-422121.gold.dim_fornecedor` f
                ON CAST(c.CNPJ_fornecedor AS STRING) = CAST(f.cnpj_completo AS STRING)
            WHERE 
                c.data_emissao BETWEEN '{dt_inicio}' AND '{dt_fim_evento}'
                AND CAST(c.SeqProduto AS STRING) IN UNNEST({skus_array_str})
                AND CAST(c.CNPJ_fornecedor AS STRING) NOT IN UNNEST({cnpjs_array_str})
            GROUP BY 1
            ORDER BY 2 DESC
            LIMIT 10
        """
        print("\n" + "="*50)
        print(">>> SQL OFENSORES:")
        print(sql_ofensores)
        print("="*50 + "\n")

        res_ofensores = bigquery_client.run_query(sql_ofensores)
        df_ofensores = [dict(row) for row in res_ofensores] if res_ofensores else []

        # --- PASSO 4: TABELA DETALHADA (META vs REALIZADO) ---
        
        # A. Busca Metas do Django (O que deveria ter sido comprado)
        # Chave do dict será tupla: (NomeAssociado, DescricaoProduto)
        distribuicoes = ItemGradeDistribuicao.objects.filter(item_grade__grade=grade).select_related('item_grade')
        tabela_final = {}
        
        for dist in distribuicoes:
            chave = (dist.associado_nome.upper().strip(), dist.item_grade.descricao_resumida)
            tabela_final[chave] = {
                'associado': dist.associado_nome,
                'produto': dist.item_grade.descricao_resumida,
                'meta': float(dist.volume_fisico), # A META QUE FALTAVA
                'aderente': 0.0,
                'fora_prazo': 0.0,
                'outros': 0.0
            }

        # B. Busca Realizado do BigQuery
        sql_detalhe = f"""
            SELECT 
                NomeAssociado,
                CAST(SeqProduto AS STRING) as sku,
                SUM(CASE 
                    WHEN CAST(CNPJ_fornecedor AS STRING) IN UNNEST({cnpjs_array_str}) AND data_emissao <= '{dt_fim}'
                    THEN QtdCompra ELSE 0 END
                ) AS qtd_aderente,
                SUM(CASE 
                    WHEN CAST(CNPJ_fornecedor AS STRING) IN UNNEST({cnpjs_array_str}) AND data_emissao > '{dt_fim}'
                    THEN QtdCompra ELSE 0 END
                ) AS qtd_fora_prazo,
                SUM(CASE 
                    WHEN CAST(CNPJ_fornecedor AS STRING) NOT IN UNNEST({cnpjs_array_str})
                    THEN QtdCompra ELSE 0 END
                ) AS qtd_outros
            FROM `singular-ray-422121.gold.obt_tb_compra_agg`
            WHERE 
                data_emissao BETWEEN '{dt_inicio}' AND '{dt_fim_evento}'
                AND CAST(SeqProduto AS STRING) IN UNNEST({skus_array_str})
            GROUP BY 1, 2
        """
        res_detalhe = bigquery_client.run_query(sql_detalhe)
        
        # C. Cruza os dados
        if res_detalhe:
            for row in res_detalhe:
                assoc = row['NomeAssociado'].upper().strip()
                sku = row['sku']
                produto_pai = skus_map.get(sku, f"SKU {sku} (Sem vínculo)")
                
                chave = (assoc, produto_pai)
                
                # Se a loja comprou algo que não tinha meta, cria a linha na hora
                if chave not in tabela_final:
                    tabela_final[chave] = {
                        'associado': row['NomeAssociado'], # Usa o nome que veio do BQ
                        'produto': produto_pai,
                        'meta': 0.0,
                        'aderente': 0.0,
                        'fora_prazo': 0.0,
                        'outros': 0.0
                    }
                
                tabela_final[chave]['aderente'] += float(row['qtd_aderente'] or 0)
                tabela_final[chave]['fora_prazo'] += float(row['qtd_fora_prazo'] or 0)
                tabela_final[chave]['outros'] += float(row['qtd_outros'] or 0)

        # Transforma em lista ordenada
        lista_tabela = list(tabela_final.values())
        lista_tabela.sort(key=lambda x: (x['associado'], x['produto']))

        # --- PASSO 5: TOTAIS POR LOJA (PARA O GRÁFICO DE BARRAS) ---
        metas_raw = ItemGradeDistribuicao.objects.filter(item_grade__grade=grade)\
            .values('associado_nome')\
            .annotate(meta_total=Sum('volume_fisico'))
        metas_dict = {m['associado_nome'].upper().strip(): float(m['meta_total']) for m in metas_raw}

        return JsonResponse({
            'status': 'ok',
            'dados_apuracao': df_apuracao,
            'dados_ofensores': df_ofensores,
            'dados_tabela': lista_tabela,
            'metas_por_loja': metas_dict,
            'periodo': {'inicio': dt_inicio, 'fim_grade': dt_fim, 'fim_evento': dt_fim_evento}
        }, safe=False)

    except Exception as e:
        print(f"Erro Dash: {e}")
        return JsonResponse({'status': 'erro', 'message': str(e)}, status=500)


class DashboardGradeView(LoginRequiredMixin, DetailView):
    model = Grade
    template_name = 'apuracao_grade/dashboard.html'
    context_object_name = 'grade'

@require_POST
def excluir_grade(request, pk):
    """
    Exclui uma Grade inteira e todos os seus vínculos (Itens, Distribuição, etc).
    """
    grade = get_object_or_404(Grade, pk=pk)
    grade_id = grade.id
    
    try:
        grade.delete()
        messages.success(request, f"Grade #{grade_id} excluída com sucesso!")
    except Exception as e:
        messages.error(request, f"Erro ao excluir a grade: {e}")
        
    return redirect('apuracao_grade:grade-list')


def api_criar_evento_modal(request):
    """
    Cria um evento via AJAX para o modal da tela de Grade.
    """
    if request.method == 'POST':
        import json
        try:
            # Tenta ler JSON (se vier via fetch body json)
            data = json.loads(request.body)
        except:
            # Fallback para POST normal
            data = request.POST

        form = EventoForm(data)
        
        if form.is_valid():
            evento = form.save()
            return JsonResponse({
                'status': 'ok',
                'id': evento.id,
                'text': str(evento) # Retorna o __str__ do model (Nome + Ano)
            })
        else:
            # Retorna erros de validação (ex: data fim < inicio)
            return JsonResponse({'status': 'erro', 'errors': form.errors}, status=400)
            
    return JsonResponse({'status': 'erro', 'message': 'Método não permitido'}, status=405)

def editar_cabecalho_grade(request, grade_id):
    grade = get_object_or_404(Grade, pk=grade_id)
    
    if request.method == 'POST':
        form = GradeHeaderForm(request.POST, instance=grade)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 1. Salva os dados normais da Grade (Datas, Obs, Evento)
                    grade_salva = form.save()

                    # 2. Atualiza os Grupos (Apaga os antigos e cria os novos)
                    grupos_str = form.cleaned_data.get('grupos_json')
                    
                    if grupos_str:
                        # Se veio dado novo, substituímos
                        grupos_list = json.loads(grupos_str)
                        
                        # Limpa os atuais
                        grade_salva.grupos.all().delete()
                        
                        # Cria os novos
                        for item in grupos_list:
                            GradeGrupo.objects.create(
                                grade=grade_salva,
                                grupo_id=item['id'],
                                grupo_nome=item['text'] # ou item['nome'] dependendo do seu JS
                            )
                            
                messages.success(request, 'Cabeçalho e Fornecedores atualizados com sucesso!')
                return redirect('apuracao_grade:grade-detalhe', pk=grade.id) # Ajuste para sua URL correta
                
            except Exception as e:
                messages.error(request, f"Erro ao salvar grupos: {e}")
    else:
        form = GradeHeaderForm(instance=grade)
    
    # --- PREPARAÇÃO PARA O TEMPLATE ---
    # Precisamos mandar os grupos atuais para o JavaScript pré-carregar no Select2
    grupos_atuais = []
    for g in grade.grupos.all():
        grupos_atuais.append({'id': g.grupo_id, 'text': g.grupo_nome})
    
    grupos_json_inicial = json.dumps(grupos_atuais)

    return render(request, 'apuracao_grade/editar_cabecalho.html', {
        'form': form, 
        'grade': grade,
        'grupos_json_inicial': grupos_json_inicial # Manda pro template
    })

# Adicione no topo
from .distribuicao_service import executar_atualizacao_distribuicao

# Adicione a função de View
def atualizar_distribuicao_view(request):
    """
    Rota para forçar a atualização da tabela de distribuição via navegador.
    """
    if not request.user.is_superuser: # Segurança básica
        return JsonResponse({'status': 'erro', 'msg': 'Apenas admin pode fazer isso.'}, status=403)

    sucesso, mensagem = executar_atualizacao_distribuicao()
    
    if sucesso:
        return JsonResponse({'status': 'ok', 'msg': mensagem})
    else:
        return JsonResponse({'status': 'erro', 'msg': mensagem}, status=500)